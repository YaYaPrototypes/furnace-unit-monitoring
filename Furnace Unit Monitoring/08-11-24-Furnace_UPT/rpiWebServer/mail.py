#!/bin/bash
#!/bin/sh
#!/usr/bin/python3

from openpyxl import *
from datetime import datetime
import time
import smtplib

SERVER = "172.16.0.8"

path_to_file =("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")

def full_night_mail():
	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date= "23-08-2023" #now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)
	#Full Night Energy Consumption Unit
	energy_full_night=sheet.cell(row = match_date,column=2).value
	yesterday_half_night=sheet.cell(row = match_date-1,column=4).value
	full_night_unit=round(((energy_full_night-yesterday_half_night)*1000000),2)
	print("Energy Meter Reading : "+ str(energy_full_night))
	print("Unit Consumption :" +str(full_night_unit))

	full_night_charge_metal=sheet.cell(row=match_date,column=6).value
	print("charge",full_night_charge_metal)
	re=str(full_night_charge_metal)
	res=float(re)
	full_night_upt=round((full_night_unit / res),2)
	print("UPT :" +str(full_night_upt))

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net"] #,"nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(full_night_unit)
	DATA1= str(full_night_charge_metal)
	DATA2= str(full_night_upt)

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES     :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"     CHARGED METAL IN TON : " +DATA1+"\n"+"     UNITS PER TON         : "+DATA2

	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("FULL NIGHT MAIL SUCCESS FULLY SENDED")
	server.quit()

def day_shift_mail():
	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date= "23-08-2023" #now.strftime("%d-%m-%Y")
#	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)

	#DAY SHift Night Energy Consumption Unit
	energy_day_shift=sheet.cell(row = match_date,column=3).value
	today_full_night=sheet.cell(row = match_date,column=2).value
	day_shift_unit=round(((energy_day_shift-today_full_night)*1000000),2)
	print("Energy Meter Reading : "+ str(energy_day_shift))
	print("Unit Consumption :" +str(day_shift_unit))
	#DAY Shift Charged Metal
	day_shift_charge_metal=sheet.cell(row=match_date,column=5).value
	print("DAY SHift Charged MEtal",day_shift_charge_metal)
	re=str(day_shift_charge_metal)
	res=float(re)
	#DAY Shift UPT
	day_shift_upt=round((day_shift_unit / res),2)
	print("DAY SHIFT UPT :" +str(day_shift_upt))

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net"] #,"nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(day_shift_unit)
	DATA1= str(day_shift_charge_metal)
	DATA2= str(day_shift_upt)

#	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES     :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"     CHARGED METAL IN TON : " +DATA1+"\n"+"     UNITS PER TON         : "+DATA2
	TEXT ="\n"+"DATE : "+current_date +"\n"+ "DAY SHIFT UPT DETAILES      :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON          : "+DATA2


	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("DAY SHIFT MAIL SUCCESS FULLY SENDED")
	server.quit()

def half_night_mail():
	workbook = load_workbook(path_to_file,data_only=True)
	sheet = workbook.active
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)

	current_date= "23-08-2023" #now.strftime("%d-%m-%Y")
#	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	r=sheet.max_row
	for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
	        for cell in row:
	                if cell.value ==(current_date):
	                        re=sheet.cell(row=cell.row, column=cell.column)
	                        print("content Row: ",cell.row)
	                        rr=cell.row
	match_date=rr
	print(match_date)

	energy_half_night=sheet.cell(row = match_date,column=4).value
	today_day_shift=sheet.cell(row = match_date,column=3).value
	half_night_unit=round(((energy_half_night-today_day_shift)*1000000),2)
	print("Energy Meter Reading : "+ str(energy_half_night))
	print("Unit Consumption :" +str(half_night_unit))

	half_night_charge_metal=sheet.cell(row=match_date,column=7).value
	print("charge",half_night_charge_metal)
	re=str(half_night_charge_metal)
	res=float(re)

	half_night_upt=round((half_night_unit / res),2)

	print("UPT :" +str(half_night_upt))

	FROM = "RPI@texmo.net"
	TO = ["ahs@texmo.net"] #,"nta@texmo.net","svu@texmo.net"]
	SUBJECT = "TEXMO INDUSTRIES FURNACE UPT "+current_date
	DATA = str(half_night_unit)
	DATA1= str(half_night_charge_metal)
	DATA2= str(half_night_upt)
#	TEXT ="\n"+"DATE : "+current_date +"\n"+ "FULL NIGHT UPT DETAILES     :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"     CHARGED METAL IN TON : " +DATA1+"\n"+"     UNITS PER TON         : "+DATA2

	TEXT ="\n"+"DATE : "+current_date +"\n"+ "HALF NIGHT UPT DETAILES     :"+"\n"+"     UNIT CONSUMPTION : "+ DATA +"\n"+"CHARGED METAL IN TON : " +DATA1+"\n"+"UNITS PER TON          : "+DATA2


	message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

	%s

	""" % (FROM, ", ".join(TO), SUBJECT, TEXT)

	server = smtplib.SMTP(SERVER)
	server.sendmail(FROM, TO, message)
	print ("HALF NIGHT MAIL SUCCESS FULLY SENDED")
	server.quit()

while True:
	print ("MAIL Timer Start")
	FULL= "02:57:05:pm"
	DAY="02:57:12:pm"
	HALF="02:57:15:pm"
	now = datetime.now()
	current_time=now.strftime("%I:%M:%S:%P")
	print("TIME : " + current_time)
	current_date=now.strftime("%d-%m-%Y")
	print("DATE : " + current_date)
	if (current_time==FULL):
	    full_night_mail()
	elif (current_time==DAY):
	    day_shift_mail()          
	elif (current_time==HALF):
	    half_night_mail()
	time.sleep(1)

