import openpyxl
path = "Furnace_UPT.xlsx" #excel file path

wb = openpyxl.load_workbook(path)
sheet = wb.active

date = "2023-07-25"
convert_date=str(date)
unorder_date = convert_date
create_list = unorder_date.split("-")
create_list.reverse()
order_list=create_list
list_to_str = ""
for i in order_list:
  list_to_str += i
order_str=list_to_str
format_date=order_str
format_date = format_date[:2] + '-' + format_date[2:]
final_format_date = format_date
final_format_date = final_format_date[:5] + '-' + final_format_date[5:]
print(final_format_date)

r=sheet.max_row

for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False):
    
    for cell in row:
        if cell.value ==(final_format_date):
            re=sheet.cell(row=cell.row, column=cell.column)
            print("content Row: ",cell.row)
            rr=cell.row
            print("result",rr)