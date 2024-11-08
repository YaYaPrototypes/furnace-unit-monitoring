from openpyxl import *
from datetime import datetime
import time


now = datetime.now()
current_time=now.strftime("%I:%M:%S:%P")
print("TIME : " + current_time)
current_date="18-08-2023" #now.strftime("%d-%m-%Y")
print("DATE : " + current_date)
path_to_file =("Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx")

workbook = load_workbook(path_to_file)
sheet = workbook.active
r=sheet.max_row
for row in sheet.iter_rows(min_row=0, min_col=0, max_row=r, max_col=1, values_only=False): 
        for cell in row:
                if cell.value ==(current_date):
                        re=sheet.cell(row=cell.row, column=cell.column)
                        print("content Row: ",cell.row)
                        rr=cell.row
match_date=rr
print(match_date)
dummy=sheet.cell(row = match_date,column=8).value
print ("this is dummy",dummy)
energy_full_night=sheet.cell(row = match_date,column=2).value
yesterday_half_night=sheet.cell(row = match_date-1,column=4).value
result=round(((energy_full_night-yesterday_half_night)*1000000),2)
#reference
result1=round(((energy_full_night-dummy)*1000000),2)
print("result1 : ",result1)
charge_metal=sheet.cell(row=match_date,column=6).value
print("charge",charge_metal)
re=str(charge_metal)
res=float(re)

upt=round((result / res),2)

print("Energy Meter Reading : "+ str(energy_full_night))
print("Unit Consumption :" +str(result))
print("UPT :" +str(upt))

