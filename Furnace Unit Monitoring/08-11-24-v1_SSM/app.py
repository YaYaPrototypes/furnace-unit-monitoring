from flask import Flask, render_template, request, redirect, url_for, jsonify, session,flash
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import traceback  # Import traceback module


app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'fallback_secret_key')
current_dir = os.path.abspath(os.path.dirname(__file__))
@app.route('/')
def index():
    file_list = get_existing_files()
    return render_template('index.html', file_list=file_list)

@app.route('/bar')
def bar():
    return render_template('barcode.html')

@app.route('/create_excel', methods=['POST'])
def create_excel():
    excel_name = request.form['excel_name'].strip()
    if not excel_name:
        return redirect(url_for('index'))
    file_path = os.path.join(current_dir, f'{excel_name}.xlsx')
    if os.path.exists(file_path):
        return redirect(url_for('view_excel', filename=excel_name))
    workbook = Workbook()
    all_spare_sheet = workbook.create_sheet(title="all_spare")
    all_spare_sheet.append(["Spare ID", "Spare Name", "Quantity", "Bero-Rack", "Booking"])
    history_sheet = workbook.create_sheet(title="history")
    history_sheet.append(["Date", "Event"])
    workbook.save(file_path)
    return redirect(url_for('view_excel', filename=excel_name))

@app.route('/view_excel/<filename>')
def view_excel(filename):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    if not os.path.exists(file_path):
        return "File not found."
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    return render_template('view_excel.html', filename=filename, sheet_names=sheet_names)

@app.route('/create_sheet/<filename>', methods=['POST'])
def create_sheet(filename):
    sheet_name = request.form['sheet_name']
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    if not os.path.exists(file_path):
        return "File not found."
    workbook = load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        return "Sheet already exists."
    workbook.create_sheet(title=sheet_name)
    sheet = workbook[sheet_name]
    title = "PM Calendar"  # You can modify the title as needed
    headers = ["Spare ID", "Spare Name", "Quantity", "Purpose"]
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value=title)
    sheet.append(headers)
    workbook.save(file_path)
    default_sheets = ['all_spare', 'history']
    for default_sheet in default_sheets:
        if default_sheet not in workbook.sheetnames:
            workbook.create_sheet(title=default_sheet)
    workbook.save(file_path)
    return redirect(url_for('view_excel', filename=filename))
def log_history_operation(wb, operation, details,quantity,bn,rn,bin,purpose):
    history_sheet = wb['history'] if 'history' in wb.sheetnames else wb.create_sheet('history')
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    history_sheet.append([current_date, operation, details,quantity,bn,rn,bin,purpose])
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(current_dir, "Furnace.xlsx"))

@app.route('/create_spare/<filename>/<sheet_name>', methods=['POST'])
def create_spare(filename, sheet_name):
    spare_id = request.form['spare_id']
    spare_name = request.form['spare_name']
    quantity = int(request.form['quantity'])
   # purpose = request.form['purpose']
    purpose = ""
    bn=request.form['bn']
    rn=request.form['rn']
    bin=request.form['bin']
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row), start=2):
        if row and len(row) > 0:  # Check if the row is not empty
            existing_spare_id = row[0].value
            if existing_spare_id == spare_id:
                existing_quantity_cell = sheet.cell(row=row_num, column=3)  # Assuming quantity is in the third column
                existing_quantity = existing_quantity_cell.value
                existing_quantity_cell.value = existing_quantity + quantity
                booking_column_index = 5  # Assuming booking is in the fifth column
               # existing_booking_cell = sheet.cell(row=row_num, column=booking_column_index)
               # existing_booking_cell.value = "YourBookingInfo"  # Replace with your actual booking information
                log_history_operation(wb, filename, f"Updated quantity and booking for Spare ID {spare_id}",quantity,bn,rn,bin,purpose)
                wb.save(file_path)
                return redirect(url_for('view_sheet', filename=filename, sheet_name='all_spare'))
    sheet.append([spare_id, spare_name, quantity,bn,rn,bin,None])
    log_history_operation(wb, filename, f"Created new spare with Spare ID {spare_id}",quantity,bn,rn,bin,purpose)
    wb.save(file_path)
    return redirect(url_for('view_sheet', filename=filename, sheet_name='all_spare'))


@app.route('/view_sheet/<filename>/<sheet_name>')
def view_sheet(filename, sheet_name):
    session['current_sheet'] = sheet_name
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    if not os.path.exists(file_path):
        return "File not found."
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[2]]
    data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        data.append(row)
    return render_template('view_sheet.html', filename=filename, sheet_name=sheet_name, headers=headers, data=data)

def get_existing_files():
    files = [f[:-5] for f in os.listdir(current_dir) if f.endswith('.xlsx')]
    return files

@app.route('/create_task/<filename>/<sheet_name>', methods=['POST'])
def create_task(filename, sheet_name):
    task_spare_id = request.form['task_spare_id']
    task_spare_name = request.form['task_spare_name']
    task_qty = request.form['task_qty']
    task_purpose = request.form['task_purpose']
    task_data = {
        'Spare ID': task_spare_id,
        'Spare Name': task_spare_name,
        'Quantity': task_qty,
        'Purpose': task_purpose,
    }
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    wb = load_workbook(file_path)
    task_sheet_name = session.get('current_sheet', 'default_sheet_name')
    sheet = wb[task_sheet_name]
    row_number = sheet.max_row + 1
    for col_num, (header, value) in enumerate(task_data.items(), start=1):
        sheet.cell(row=row_number, column=col_num, value=value)
    wb.save(file_path)
    update_all_spare_sheet(wb, task_spare_id, task_qty)
    log_history_operation(wb, 'Task Created', f'Task for Spare ID {task_spare_id} created with Quantity {task_qty}')
    return redirect(url_for('view_sheet', filename=filename, sheet_name=task_sheet_name))

@app.route('/Spare_remove/<filename>/<sheet_name>', methods=['POST'])
def spare_remove(filename, sheet_name):
    spare_id = request.form['spare_id']
    spare_name = request.form['spare_name']
    quantity = int(request.form['quantity'])
    purpose = request.form['purpose']
    bn=request.form['bn']
    rn=request.form['rn']
    bin=request.form['bin']
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row), start=2):
        if row and len(row) > 0:  # Check if the row is not empty
            existing_spare_id = row[0].value
            if existing_spare_id == spare_id:
                existing_quantity_cell = sheet.cell(row=row_num, column=3)  # Assuming quantity is in the third column
                existing_quantity = existing_quantity_cell.value
                existing_quantity_cell.value = existing_quantity - quantity
                print("this is ",existing_quantity_cell.value)
                booking_column_index = 5  # Assuming booking is in the fifth column
               # existing_booking_cell = sheet.cell(row=row_num, column=booking_column_index)
               # existing_booking_cell.value = "YourBookingInfo"  # Replace with your actual booking information
                log_history_operation(wb, filename, f"spare Used  with this Spare ID:  {spare_id}",quantity,bn,rn,bin,str(purpose))
                wb.save(file_path)
                return redirect(url_for('view_sheet', filename=filename, sheet_name='all_spare'))
    sheet.append([spare_id, spare_name, quantity, purpose,None])
    log_history_operation(wb, filename, f"spare Used  with this Spare ID:  {spare_id}",quantity,bn,rn,bin,purpose)
    wb.save(file_path)
    return redirect(url_for('view_sheet', filename=filename, sheet_name='all_spare'))



def update_all_spare_sheet(wb, task_spare_id, task_qty):
    all_spare_sheet = wb['all_spare']
    for row in all_spare_sheet.iter_rows(min_row=2, max_col=5, max_row=all_spare_sheet.max_row):
        existing_spare_id, existing_spare_name, existing_qty, purpose, existing_booking_info = [cell.value for cell in row]
        if existing_spare_id == task_spare_id:
            try:
                if existing_booking_info is not None:
                    updated_booking_info = str(int(existing_booking_info) + int(task_qty))
                else:
                    updated_booking_info = str(task_qty)
                row[4].value = updated_booking_info
            except ValueError:
                print(f"Unable to update booking information for Spare ID {task_spare_id}. Skipping update.")
            break  # Stop searching once the spare ID is found and updated
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(current_dir, 'Furnace.xlsx'))

@app.route('/get_spare_name/<filename>/<spare_id>')
def get_spare_name(filename, spare_id):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    if not os.path.exists(file_path):
        return jsonify({"success": False, "spare_name": None})
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook['all_spare'] if 'all_spare' in workbook.sheetnames else None
    if sheet:
        for row in sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row):
            existing_spare_id, existing_spare_name ,existing_qty , existing_bero ,existing_rack , existing_bin = row[0].value, row[1].value,row[2].value,row[3].value,row[4].value,row[5].value
            if existing_spare_id == spare_id:
                return jsonify({"success": True , "spare_name": existing_spare_name , "quantity": existing_qty , "bn": existing_bero , "rn": existing_rack , "bin": existing_bin})
    return jsonify({"success": False, "spare_name": None})

@app.route('/get_spare_name1/<filename>/<spare_id>')
def get_spare_name1(filename, spare_id):
    file_path = os.path.join(current_dir, f'{filename}.xlsx')
    if not os.path.exists(file_path):
        return jsonify({"success": False, "spare_name1": None})
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook['all_spare'] if 'all_spare' in workbook.sheetnames else None
    if sheet:
        for row in sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row):
            existing_spare_id1, existing_spare_name1 ,existing_qty1 , existing_bero1 ,existing_rack1 , existing_bin1 = row[0].value, row[1].value,row[2].value,row[3].value,row[4].value,row[5].value
            if existing_spare_id1 == spare_id:
                return jsonify({"success": True , "spare_name1": existing_spare_name1 , "quantity1": existing_qty1 , "bn1": existing_bero1 , "rn1": existing_rack1 , "bin1": existing_bin1})
    return jsonify({"success": False, "spare_name1": None})




#        for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row):
#            existing_spare_id1, existing_spare_name1 = row[0].value, row[1].value
#            if existing_spare_id1 == spare_id:
#                return jsonify({"success": True, "spare_name1": existing_spare_name1})
#    return jsonify({"success": False, "spare_name1": None})



def update_all_spare_after_completion(wb, task_spare_id, task_qty):
    all_spare_sheet = wb['all_spare']
    for row in all_spare_sheet.iter_rows(min_row=2, max_col=5, max_row=all_spare_sheet.max_row):
        existing_spare_id, existing_spare_name, existing_qty, purpose, existing_booking_info = [cell.value for cell in row]
        if existing_spare_id == task_spare_id:
            try:
                updated_qty = int(existing_qty) - int(task_qty)
                updated_qty = max(updated_qty, 0)
                row[2].value = updated_qty
            except (ValueError, TypeError) as e:
                print(f"Unable to update quantity for Spare ID {task_spare_id}. Skipping update. Error: {e}")
            break  # Stop searching once the spare ID is found and updated
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb.save(os.path.join(current_dir, 'Furnace.xlsx'))

@app.route('/complete_task/<filename>/<sheet_name>/<int:row_number>', methods=['POST'])
def complete_task(filename, sheet_name, row_number):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    wb = load_workbook(os.path.join(current_dir, f'{filename}.xlsx'))
    sheet = wb[sheet_name]
    task_spare_id = sheet.cell(row=row_number, column=1).value
    task_qty = sheet.cell(row=row_number, column=3).value
    update_all_spare_after_completion(wb, task_spare_id, task_qty)
    sheet.delete_rows(row_number)
    wb.save(os.path.join(current_dir, f'{filename}.xlsx'))
    return jsonify({"success": True, "message": "Task marked as complete."})



if __name__ == '__main__':
#    app.run(host='172.16.5.190',port =5000 ,debug=True)
    app.run()
