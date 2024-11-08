from flask import Flask, jsonify, render_template, request
from pymodbus.client.sync import ModbusTcpClient
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from datetime import datetime
import openpyxl
import os

file_name="/home/store/Furnace_Energy_Units_Monitor/v2.0/Data.xlsx"
os.system(f'sudo chown pi:pi {file_name}')
sheet_name ="Sheet1"
headers = ["Date","Full Night","Day Shift","Half Night","Full Night Units","Day Shift Units","Half Night Units"]
if not os.path.exists(file_name):
    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
else:
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

# Write headers to the worksheet
workbook.save(file_name)

app = Flask(__name__)

client = ModbusTcpClient("172.16.4.235", port=502)

#logging times 
full="07:30:00" #07:30:00
day ="17:30:00" #17:30:00
half="21:30:00" #21:30:00

full_set=["21","22","23","24","00","01","02","03","04","05","06"]
day_set=["07","08","09","10","11","12","13","14","15","16"]
half_set=["17","18","19","20"]

full_time = datetime.strptime(full, "%H:%M:%S").time()
day_time = datetime.strptime(day, "%H:%M:%S").time()
half_time = datetime.strptime(half, "%H:%M:%S").time()

def full_log(current_kwh,current_date):
    print("yes i'm working","FULL_night Function",current_kwh,current_date)
    # curret_date = datetime.now().strftime("%Y-%m-%d")
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]  
    print("pass")
    max_row=0
    for cell in worksheet["B"]:
        if cell.value is not None:
            max_row = cell.row
    print(max_row)

    next_row = max_row + 1
    worksheet.cell(row=next_row, column=2, value=current_kwh)
    worksheet.cell(row=next_row, column=1, value=current_date)
    worksheet.cell(row=next_row, column=5, value=f'=(B{next_row}-D{next_row-1})*1000000')
    worksheet.cell(row=next_row, column=11, value=f'=(E{next_row}/H{next_row})')

    workbook.save(file_name)


def day_log(current_kwh,current_date):
    print("yes i'm working","Day Shift Function",current_kwh)
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]  
    print("pass")
    max_row=0
    for cell in worksheet["C"]:
        if cell.value is not None:
            max_row = cell.row
    print(max_row)

    next_row = max_row + 1
    worksheet.cell(row=next_row, column=3, value=current_kwh)
    worksheet.cell(row=next_row, column=6, value=f'=(C{next_row}-B{next_row})*1000000')
    worksheet.cell(row=next_row, column=12, value=f'=(F{next_row}/I{next_row})')

    workbook.save(file_name)

def half_log(current_kwh,current_date):
    print("yes i'm working","Hlaf_night Function",current_kwh)
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name] 
    print("pass")
    max_row=0
    for cell in worksheet["D"]:
        if cell.value is not None:
            max_row = cell.row
    print(max_row)

    next_row = max_row + 1
    worksheet.cell(row=next_row, column=4, value=current_kwh)
    worksheet.cell(row=next_row, column=7, value=f'=(D{next_row}-C{next_row})*1000000')
    worksheet.cell(row=next_row, column=13, value=f'=(G{next_row}/J{next_row})')

    workbook.save(file_name)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_data')
def get_data():
    try:
        client.connect()
        data = client.read_holding_registers(257, 2, slave=1)
        # shift = None
        if not data.isError():
            decoder = BinaryPayloadDecoder.fromRegisters(data.registers, Endian.Big, wordorder=Endian.Little)
            volt = decoder.decode_32bit_float()
            vo = volt / 1000000000
            current_kwh = round(vo, 6)
            
            # Get the current time
            current_time = datetime.now().strftime('%H:%M:%S')
            h=datetime.now().strftime("%H")
            hour=str(h)
            print(hour)
            current_date = datetime.now().strftime("%Y-%m-%d")
            if current_time ==   full:
                #shift= "full"
                print("Activate full night kwh reading","kwh : ",current_kwh)
                full_log(current_kwh,current_date)
            elif current_time == day:
                # shift= "day"
                print("Activate day shift kwh reading","kwh : ",current_kwh)
                day_log(current_kwh,current_date)
            elif current_time == half:
                # shift= "half"
                print("Activate half night kwh reading","kwh : ",current_kwh)
                half_log(current_kwh,current_date)
        # Live Shift Update

            if hour in full_set:  # Full Night Shift is from 9:30 PM to 4:30 AM
                shift = "Full Night Shift"
            elif hour in day_set:  # Day Shift is from 7:30 AM to 5:30 PM
                shift = "Day Shift"
            elif hour in half_set:  # Half Night Shift is from 4:30 AM to 7:30 AM
                shift = "Half Night Shift"
            else:
                shift = "Unknown Shift"  # If somehow the time doesn't fall in any range
            print ("current shift get data",shift)


        #refernce Kwh
            if shift == "Full Night Shift":
                max_row=0
                for cell in worksheet["D"]:
                    if cell.value is not None:
                        max_row = cell.row
                print("Full Night Shift",max_row)
                ref_kwh= worksheet.cell(row=max_row, column=4).value
                
            elif shift == "Day Shift":
                max_row=0
                for cell in worksheet["C"]:
                    if cell.value is not None:
                        max_row = cell.row
                print("Day Shift",max_row)
                ref_kwh= worksheet.cell(row=max_row, column=2).value
            elif shift == "Half Night Shift":
                max_row=0
                for cell in worksheet["B"]:
                    if cell.value is not None:
                        max_row = cell.row
                print("Half Night Shift",max_row)
                ref_kwh= worksheet.cell(row=max_row, column=3).value   
            
            print("result",ref_kwh)      


            return jsonify({'voltage': current_kwh, 'time': current_time,'date':current_date,'shift':shift,'ref_kwh':ref_kwh})
        else:
            return jsonify({'error': 'Error reading data'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        client.close()

@app.route('/ref_panel')
def ref_panel():
    max_row = worksheet.max_row

    # Retrieve values from the worksheet and handle None values by setting them to 0
    ref_kwh_half_get = worksheet.cell(row=max_row, column=4).value or 0
    ref_kwh_full_get = worksheet.cell(row=max_row, column=2).value or 0
    ref_kwh_day_get = worksheet.cell(row=max_row, column=3).value or 0
    ref_kwh_full_in_get = worksheet.cell(row=max_row-1, column=4).value or 0

    #UPT
    upt_full = worksheet.cell(row=max_row, column=11).value or 0
    upt_day = worksheet.cell(row=max_row, column=12).value or 0
    upt_half = worksheet.cell(row=max_row, column=13).value or 0
    # round with dismiss point values
    # upt_full = round(upt_full) if upt_full != 0 else 0
    # upt_day = round(upt_day) if upt_day != 0 else 0
    # upt_half = round(upt_half) if upt_half != 0 else 0
    print("full",ref_kwh_full_get,"day", ref_kwh_day_get, "half",ref_kwh_half_get,"full ini",ref_kwh_full_in_get)
    # Print the values retrieved
    print("upt_full",upt_full,"upt_day", upt_day, "upt_half",upt_half)

    # Perform subtraction and ensure no negative values
    ref_kwh_full = max(0,(ref_kwh_full_get - ref_kwh_full_in_get)*1000000)
    ref_kwh_day =  max(0,(ref_kwh_half_get - ref_kwh_full_get)*1000000)
    ref_kwh_half = max(0,(ref_kwh_half_get - ref_kwh_day_get)*1000000)

    # ref_kwh_full = max(0, ref_kwh_full_get - ref_kwh_full_in_get)
    # ref_kwh_day = max(0, ref_kwh_half_get - ref_kwh_full_get)
    # ref_kwh_half = max(0, ref_kwh_full_get - ref_kwh_day_get)

    ref_kwh_full = round(ref_kwh_full) if ref_kwh_full != 0 else 0
    ref_kwh_day = round(ref_kwh_day) if ref_kwh_day != 0 else 0
    ref_kwh_half = round(ref_kwh_half) if ref_kwh_half != 0 else 0

    # Print the results
    print(ref_kwh_full, ref_kwh_day, ref_kwh_half)
    return jsonify({'ref_kwh_full':ref_kwh_full,'ref_kwh_day':ref_kwh_day,'ref_kwh_half':ref_kwh_half,'upt_full':upt_full,'upt_day':upt_day,'upt_half':upt_half})



@app.route('/submit_form', methods=['POST'])
def submit_form():
    try:
        # Load the workbook and sheet
        workbook = openpyxl.load_workbook(file_name ,data_only=True)
        worksheet = workbook[sheet_name] 

        # Get the current time
        current_time = datetime.now().strftime('%H:%M:%S')
        h=datetime.now().strftime("%H")
        hour=str(h)
        print(hour)
        current_time_obj = datetime.strptime(current_time, "%H:%M:%S").time()

        print("Current Time:", current_time)

        # Get data from the form
        data = request.json
        print("Form data received:", data)
        
        # Get 'ton' value from the form
        ton_value = data.get('ton')
        
        print("Ton value:", ton_value)

        # Check if 'ton' value is None
        if ton_value is None:
            return jsonify({'error': 'Missing ton data'}), 400

        # Determine which shift the current time falls into
        if hour in full_set:  # Full Night Shift is from 9:30 PM to 4:30 AM
            shift = "Full Night Shift"
        elif hour in day_set:  # Day Shift is from 7:30 AM to 5:30 PM
            shift = "Day Shift"
        elif hour in half_set:  # Half Night Shift is from 4:30 AM to 7:30 AM
            shift = "Half Night Shift"
        else:
            shift = "Unknown Shift"  # If somehow the time doesn't fall in any range
        print ("current shift get data",shift)

        if shift == "Full Night Shift":
                print("pass FULL NIGHT")
                # max_row=0
                # for cell in worksheet["H"]:
                #     if cell.value is not None:
                #         max_row = cell.row
                # print(max_row)
                next_row = worksheet.max_row #+ 1
                worksheet.cell(row=next_row, column=10, value=ton_value)
                workbook.save(file_name)
        elif shift == "Day Shift":
                print("pass DAY")
                # max_row=0
                # for cell in worksheet["I"]:
                #     if cell.value is not None:
                #         max_row = cell.row
                # print(max_row)
                next_row = worksheet.max_row   #+ 1
                worksheet.cell(row=next_row, column=8, value=ton_value)
                workbook.save(file_name)
        elif shift == "Half Night Shift":
                print("pass HALF NIGHT")
                # max_row=0
                # for cell in worksheet["J"]:
                #     if cell.value is not None:
                #         max_row = cell.row
                # print(max_row)
                next_row = worksheet.max_row  #+ 1
                worksheet.cell(row=next_row, column=9, value=ton_value)
                workbook.save(file_name)


        # Return the shift information along with received data
        return jsonify({
            'message': 'Form data received',
            'shift': shift,
            'data': data
        }), 200

    except Exception as e:
        print("Error:", str(e))  # Print the error for debugging
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='172.16.5.190',port =5003 ,debug=True)
