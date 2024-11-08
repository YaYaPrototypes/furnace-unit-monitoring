from flask import Flask, jsonify, render_template, request
from pymodbus.client.sync import ModbusTcpClient
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from datetime import datetime
import openpyxl
import os

# File path and sheet initialization
file_name = "/home/store/Furnace_Energy_Units_Monitor/v2.0/Data.xlsx"
sheet_name = "Sheet1"
headers = ["Date", "Full Night", "Day Shift", "Half Night", "Full Night Units", "Day Shift Units", "Half Night Units"]

# Check if file exists and create if not
if not os.path.exists(file_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
else:
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

# Save the workbook after checking
workbook.save(file_name)

# Initialize Flask app
app = Flask(__name__)

# Modbus Client Setup
client = ModbusTcpClient("172.16.4.235", port=502)

# Timing for shifts
full = "21:30:00"  # Full night shift starts at 9:30 PM
day = "07:30:00"   # Day shift starts at 7:30 AM
half = "17:30:00"  # Half night shift starts at 5:30 PM

full_time = datetime.strptime(full, "%H:%M:%S").time()
day_time = datetime.strptime(day, "%H:%M:%S").time()
half_time = datetime.strptime(half, "%H:%M:%S").time()
@app.route('/')
def index():
    return render_template('index.html')

# Function to log shift data into the Excel sheet
def log_data_for_shift(shift, current_kwh, current_date):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    
    shift_column_map = {
        "Full Night": 2,
        "Day": 3,
        "Half Night": 4
    }
    
    shift_column = shift_column_map.get(shift)
    if not shift_column:
        print("Invalid shift")
        return

    max_row = worksheet.max_row

    # Check if data already exists for this shift on this date
    shift_logged = False
    for row in range(2, max_row + 1):
        if worksheet.cell(row=row, column=1).value == current_date and worksheet.cell(row=row, column=shift_column).value is not None:
            shift_logged = True
            break

    if shift_logged:
        print(f"Data already logged for {shift} on {current_date}. Skipping.")
        return

    next_row = max_row + 1
    worksheet.cell(row=next_row, column=1, value=current_date)  # Date
    worksheet.cell(row=next_row, column=shift_column, value=current_kwh)  # Shift value (Full Night, Day, Half Night)

    # Calculate units and write formulas
    if shift == "Full Night":
        worksheet.cell(row=next_row, column=5, value=f'=IFERROR((B{next_row}-D{next_row-1})*1000000, 0)')
        worksheet.cell(row=next_row, column=11, value=f'=IFERROR(E{next_row}/H{next_row}, 0)')
    elif shift == "Day":
        worksheet.cell(row=next_row, column=6, value=f'=IFERROR((C{next_row}-B{next_row})*1000000, 0)')
        worksheet.cell(row=next_row, column=12, value=f'=IFERROR(F{next_row}/I{next_row}, 0)')
    elif shift == "Half Night":
        worksheet.cell(row=next_row, column=7, value=f'=IFERROR((D{next_row}-C{next_row})*1000000, 0)')
        worksheet.cell(row=next_row, column=13, value=f'=IFERROR(G{next_row}/J{next_row}, 0)')

    # Save the changes to the Excel file
    workbook.save(file_name)
    print(f"{shift} data logged for {current_date}.")

# Function to determine the current shift
def get_shift():
    current_time = datetime.now().strftime('%H:%M:%S')
    current_time_obj = datetime.strptime(current_time, "%H:%M:%S").time()
    current_date = datetime.now().strftime("%Y-%m-%d")

    if full_time <= current_time_obj < day_time:  # Full Night Shift
        shift = "Full Night"
    elif day_time <= current_time_obj < half_time:  # Day Shift
        shift = "Day"
    elif half_time <= current_time_obj or current_time_obj < full_time:  # Half Night Shift
        shift = "Half Night"
    else:
        shift = "Unknown Shift"
    
    return shift, current_date

# API to read Modbus data and log it
@app.route('/get_data')
def get_data():
    try:
        client.connect()
        data = client.read_holding_registers(257, 2, slave=1)
        if not data.isError():
            decoder = BinaryPayloadDecoder.fromRegisters(data.registers, Endian.Big, wordorder=Endian.Little)
            volt = decoder.decode_32bit_float()
            vo = volt / 1000000000
            current_kwh = round(vo, 6)

            shift, current_date = get_shift()

            # Log the data for the specific shift
            log_data_for_shift(shift, current_kwh, current_date)

            return jsonify({'voltage': current_kwh, 'time': datetime.now().strftime('%H:%M:%S'), 'date': current_date, 'shift': shift})

        return jsonify({'error': 'Error reading data'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        client.close()

# API to display reference data
@app.route('/ref_panel')
def ref_panel():
    max_row = worksheet.max_row

    ref_kwh_half_get = worksheet.cell(row=max_row, column=4).value or 0
    ref_kwh_full_get = worksheet.cell(row=max_row, column=2).value or 0
    ref_kwh_day_get = worksheet.cell(row=max_row, column=3).value or 0
    ref_kwh_full_in_get = worksheet.cell(row=max_row - 1, column=4).value or 0

    # UPT values
    upt_full = worksheet.cell(row=max_row, column=11).value or 0
    upt_day = worksheet.cell(row=max_row, column=12).value or 0
    upt_half = worksheet.cell(row=max_row, column=13).value or 0

    # Perform the calculations
    ref_kwh_full = max(0, (ref_kwh_full_get - ref_kwh_full_in_get) * 1000000)
    ref_kwh_day = max(0, (ref_kwh_half_get - ref_kwh_full_get) * 1000000)
    ref_kwh_half = max(0, (ref_kwh_half_get - ref_kwh_day_get) * 1000000)

    ref_kwh_full = round(ref_kwh_full) if ref_kwh_full != 0 else 0
    ref_kwh_day = round(ref_kwh_day) if ref_kwh_day != 0 else 0
    ref_kwh_half = round(ref_kwh_half) if ref_kwh_half != 0 else 0

    return jsonify({'ref_kwh_full': ref_kwh_full, 'ref_kwh_day': ref_kwh_day, 'ref_kwh_half': ref_kwh_half, 'upt_full': upt_full, 'upt_day': upt_day, 'upt_half': upt_half})

# Submit form API for manual data input
@app.route('/submit_form', methods=['POST'])
def submit_form():
    try:
        workbook = openpyxl.load_workbook(file_name, data_only=True)
        worksheet = workbook[sheet_name]

        current_time = datetime.now().strftime('%H:%M:%S')
        current_time_obj = datetime.strptime(current_time, "%H:%M:%S").time()
        current_date = datetime.now().strftime("%Y-%m-%d")

        data = request.json
        ton_value = data.get('ton')

        if ton_value is None:
            return jsonify({'error': 'Missing ton data'}), 400

        shift = ""
        if current_time_obj >= full_time and current_time_obj < day_time:
            shift = "FULL NIGHT"
        elif current_time_obj >= day_time and current_time_obj < half_time:
            shift = "DAY"
        elif current_time_obj >= half_time or current_time_obj < full_time:
            shift = "HALF NIGHT"

        log_data_for_shift(shift, ton_value, current_date)

        return jsonify({
            'message': 'Form data received',
            'shift': shift,
            'data': data
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='172.16.5.190', port=5003, debug=True)
