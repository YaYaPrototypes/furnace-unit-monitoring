from openpyxl import Workbook, load_workbook

# Load or create the workbook
file = "sample.xlsx"
try:
    wb = load_workbook(file)
except FileNotFoundError:
    wb = Workbook()

# Input for the sheet name
sheet_name = input("Enter the sheet name: ")
ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

# Create header if the sheet is new
if ws.max_row == 1 and ws['A1'].value is None:
    ws['A1'] = "User"  

def get_max_row_in_column(column_letter):
    # Get the maximum row number that is not empty in the specified column
    max_row = 0
    for cell in ws[column_letter]:
        if cell.value is not None:
            max_row = cell.row
    return max_row

# Input for the specific column to check
column_letter = input("Enter the column letter to check (e.g., 'A'): ").upper()
max_row = get_max_row_in_column(column_letter)
print(f"The maximum row with data in column {column_letter} is: {max_row}")

process = input("Enter process number (1 to write, 2 to delete): ")
if process == "1":
    user = input("Enter user: ")
    ws.append([user])  # Append the user to the next available row in the first column
    wb.save(file)
elif process == "2":
    # Implement delete functionality if needed
    print("Delete functionality not implemented.")
else:
    print("Process is not available")
