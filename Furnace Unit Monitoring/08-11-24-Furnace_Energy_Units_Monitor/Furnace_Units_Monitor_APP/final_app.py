from flask import Flask, render_template, jsonify
from datetime import datetime
import openpyxl
import re
import os
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.client.sync import ModbusTcpClient as ModbusClient

app = Flask(__name__)

@app.route('/')
def cunit():
    # Read shift timings directly in this function
    time_ranges = {}
    shifts_file_path = 'shifts.txt'  # Update this path as needed

    with open(shifts_file_path, 'r') as file:
        for line in file:
            raw_line = line.strip()
            match = re.match(r"(\w+)=start:(\d{2}:\d{2}) end:(\d{2}:\d{2})", raw_line)
            if match:
                shift_name = match.group(1)
                start_time_str = match.group(2)
                end_time_str = match.group(3)

                start_time = datetime.strptime(start_time_str, "%H:%M").time()
                end_time = datetime.strptime(end_time_str, "%H:%M").time()
                time_ranges[shift_name] = (start_time, end_time)

    path_to_file = '/home/pi/20-10-23-Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx'  # Change to your actual Excel file path
    workbook = openpyxl.load_workbook(path_to_file)
    sheet = workbook.active
    current_shift_value = None
    current_shift_name = None
    current_time = datetime.now().time()

    for shift_name, (start, end) in time_ranges.items():
        if (start <= end and start <= current_time <= end) or (start > end and (current_time >= start or current_time <= end)):
            current_shift_name = shift_name
            break

    if current_shift_name:
        max_value = None
        column_mapping = {
            'fullnight': 'B',
            'dayshift': 'C',
            'halfnight': 'D'
        }
        current_column = column_mapping.get(current_shift_name)

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f'{current_column}{row}'].value
            if isinstance(cell_value, (int, float)):
                if max_value is None or cell_value > max_value:
                    max_value = cell_value

        current_shift_value = max_value

    return render_template('cunit.html', current_shift_name=current_shift_name)

@app.route('/update_data_cunit')
def update_data_cunit():
    # Read from Modbus
    data = client.read_holding_registers(HMI_Doller_Address, 2, slave=1)
    kwh_data = None
    current_shift_value = None
    
    if not data.isError():
        decoder = BinaryPayloadDecoder.fromRegisters(data.registers, Endian.Big, wordorder=Endian.Little)
        address_result = decoder.decode_32bit_float()
        kwh_data = round(address_result / 1000000000, 14)

    # Read shift timings directly in this function
    time_ranges = {}
    shifts_file_path = 'shifts.txt'  # Update this path as needed

    with open(shifts_file_path, 'r') as file:
        for line in file:
            raw_line = line.strip()
            match = re.match(r"(\w+)=start:(\d{2}:\d{2}) end:(\d{2}:\d{2})", raw_line)
            if match:
                shift_name = match.group(1)
                start_time_str = match.group(2)
                end_time_str = match.group(3)

                start_time = datetime.strptime(start_time_str, "%H:%M").time()
                end_time = datetime.strptime(end_time_str, "%H:%M").time()
                time_ranges[shift_name] = (start_time, end_time)

    path_to_file = '/home/pi/20-10-23-Furnace_UPT/rpiWebServer/Furnace_UPT.xlsx'  # Change to your actual Excel file path
    workbook = openpyxl.load_workbook(path_to_file)
    sheet = workbook.active
    current_time = datetime.now().time()

    for shift_name, (start, end) in time_ranges.items():
        if (start <= end and start <= current_time <= end) or (start > end and (current_time >= start or current_time <= end)):
            current_shift_name = shift_name
            break
    
    if current_shift_name:
        max_value = None
        column_mapping = {
            'fullnight': 'D',
            'dayshift': 'B',
            'halfnight': 'C'
        }
        current_column = column_mapping.get(current_shift_name)
        print(current_column)
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f'{current_column}{row}'].value
            if isinstance(cell_value, (int, float)):
                if max_value is None or cell_value > max_value:
                    max_value = cell_value

        current_shift_value = max_value

    # Calculate the difference
    difference = None
    if kwh_data is not None and current_shift_value is not None:
        print(((kwh_data - current_shift_value)*1000000))
        difference = int(((kwh_data - current_shift_value)*1000000))
    print(current_shift_name)
    return jsonify(kwh_data=kwh_data, current_shift_value=current_shift_value, difference=difference,current_shift_name=current_shift_name)

if __name__ == '__main__':
    ip_address = "172.16.4.235"
    HMI_Doller_Address = 257
    client = ModbusClient(ip_address, port=502)
    client.connect()
    app.run(host='172.16.5.190',port =5003 ,debug=True)
