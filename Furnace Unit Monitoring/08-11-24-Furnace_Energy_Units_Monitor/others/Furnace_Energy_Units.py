from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.client.sync import ModbusTcpClient
import time
import openpyxl
import os
from datetime import datetime
# Initialize Modbus TCP client
client = ModbusTcpClient("172.16.4.235", port=502)

# Excel file path
excel_file = 'Furnace_UPT_Data.xlsx'

# Initialize the stored values
full_night_shift_value = None
day_shift_value = None
half_night_shift_value = None

# Check if the file exists; if not, create it
if not os.path.exists(excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Energy Data"
    sheet.append(["ID", "Date", "FULL NIGHT", "DAY SHIFT", "HALF NIGHT", "FULL UNITS", "DAY UNITS", "HALF UNITS"])  # Header
    workbook.save(excel_file)
else:
    # Load the workbook to retrieve the stored values
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    current_date = datetime.now().strftime("%d/%m/%y")  # Format as DD/MM/YY
    
    # Search for today's values
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == current_date:
            full_night_shift_value = row[2]  # Store FULL value
            day_shift_value = row[3]   # Store DAY value
            half_night_shift_value = row[4]  # Store HALF value
            break

# Initialize ID counter
entry_id = len(sheet['A'])  # Set ID to the next available entry

# Define target times
FULL_NIGHT_SHIFT_TIME = "03:12:10 AM"
DAY_SHIFT_TIME        = "03:12:15 AM"
HALF_NIGHT_SHIFT_TIME = "03:12:20 AM"

try:
    client.connect()
    print("OK")
    while True:
        # Read registers from the Modbus server
        data = client.read_holding_registers(257, 2, slave=1)

        if data.isError():
            print("Error")
            print(data)
        else:
            decoder = BinaryPayloadDecoder.fromRegisters(data.registers, Endian.Big, wordorder=Endian.Little)
            Energy_Meter_Reading = decoder.decode_32bit_float()
            Energy_Reading = (Energy_Meter_Reading / 1000000000)
            Energy_Reading_Final_Value = round(Energy_Reading, 6)
            print(Energy_Reading_Final_Value)
            # Get current time and date
            current_time = datetime.now().strftime("%I:%M:%S %p")
            current_date = datetime.now().strftime("%d/%m/%y")  # Format as DD/MM/YY

            # Load the workbook to check for existing date
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # Check if the date already exists in the sheet
            row_to_update = None
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[1] == current_date:
                    row_to_update = row[0]  # Store ID of the row to update
                    break

            # If date exists, find the row index
            if row_to_update:
                row_index = row_to_update + 1  # +1 for the header row

                # Update the relevant column for FULL, DAY, or HALF
                if current_time == FULL_NIGHT_SHIFT_TIME:
                    full_night_shift_value = Energy_Reading_Final_Value  # Update stored FULL value
                    sheet.cell(row=row_index, column=3, value=full_night_shift_value)  # Update FULL
                if current_time == DAY_SHIFT_TIME:
                    day_shift_value = Energy_Reading_Final_Value  # Update stored DAY value
                    sheet.cell(row=row_index, column=4, value=day_shift_value)  # Update DAY
                if current_time == HALF_NIGHT_SHIFT_TIME:
                    half_night_shift_value = Energy_Reading_Final_Value  # Update stored HALF value
                    sheet.cell(row=row_index, column=5, value=half_night_shift_value)  # Update HALF
            else:
                # If the date doesn't exist, append a new row with formulas
                new_row_index = entry_id + 1
                sheet.append([
                    entry_id, current_date,
                    Energy_Reading_Final_Value if current_time == FULL_NIGHT_SHIFT_TIME else None,
                    Energy_Reading_Final_Value if current_time == DAY_SHIFT_TIME else None,
                    Energy_Reading_Final_Value if current_time == HALF_NIGHT_SHIFT_TIME else None,
                    f'=D{new_row_index}-C{new_row_index}/100',  # FULL UNITS formula
                    f'=E{new_row_index}-D{new_row_index}/100',  # DAY UNITS formula
                    f'=C{new_row_index+1}-E{new_row_index}'  # HALF UNITS formula
                ])

                # Increment the ID for the next entry
                entry_id += 1

            # Save the workbook
            workbook.save(excel_file)

        time.sleep(1)

except Exception as e:
    print("NOT OK", e)
